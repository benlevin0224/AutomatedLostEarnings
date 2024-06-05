from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import tkinter as tk
from tkinter.filedialog import askopenfilename, askdirectory


# Function to run the lost earnings calc
def lostearnings():
    path = askopenfilename()
    wb_obj = openpyxl.load_workbook(path)
    sh = wb_obj.active

    url = "https://www.askebsa.dol.gov/vfcpcalculator/webcalculator.aspx"

    driver = webdriver.Firefox()
    driver.get(url)
    assert "VFCP Calculator" in driver.title

    for i in range(2, sh.max_row + 1):
        print("\n")
        principal = sh.cell(row=i, column=2).value
        loss_date = sh.cell(row=i, column=3).value
        recovery_date = sh.cell(row=i, column=4).value
        final_payment_date = sh.cell(row=i, column=5).value
        driver.find_element(By.ID, "_ctl0_MainContent_txtPrincipal").send_keys(principal)
        driver.find_element(By.ID, "_ctl0_MainContent_txtLossDateMonth").send_keys(loss_date.month)
        driver.find_element(By.ID, "_ctl0_MainContent_txtLossDateDay").send_keys(loss_date.day)
        driver.find_element(By.ID, "_ctl0_MainContent_txtLossDateYear").send_keys(loss_date.year)
        driver.find_element(By.ID, "_ctl0_MainContent_txtRecoveryDateMonth").send_keys(recovery_date.month)
        driver.find_element(By.ID, "_ctl0_MainContent_txtRecoveryDateDay").send_keys(recovery_date.day)
        driver.find_element(By.ID, "_ctl0_MainContent_txtRecoveryDateYear").send_keys(recovery_date.year)
        driver.find_element(By.ID, "_ctl0_MainContent_txtFinalPaymentMonth").send_keys(final_payment_date.month)
        driver.find_element(By.ID, "_ctl0_MainContent_txtFinalPaymentDay").send_keys(final_payment_date.day)
        driver.find_element(By.ID, "_ctl0_MainContent_txtFinalPaymentYear").send_keys(final_payment_date.year)
        driver.find_element(By.ID, "_ctl0_MainContent_cmdCalculate").click()

    driver.find_element(By.ID, "_ctl0_MainContent_cmdResults").click()
    # driver.save_screenshot('/Users/benlevin/Desktop/test.png')

    # driver.close()


# Function to generate new template that can be used for lost earnings
def template():
    path = askdirectory()
    filepath = path + "/Lost Earnings Template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws["A1"] = "Name"
    ws["B1"] = "Principal"
    ws["C1"] = "Loss Date"
    ws["D1"] = "Recovery Date"
    ws["E1"] = "Final Payment Date"
    wb.save(filepath)


window = tk.Tk()

window.geometry("400x400")
window.title("Automated Lost Earnings")

btn01 = tk.Button(
    window,
    text="Run Lost Earnings",
    command=lostearnings,
)

btn02 = tk.Button(
    window,
    text="Create Template",
    command=template,
)

btn01.place(
    relx=0.5,
    rely=0.5,
    anchor="center"
)

btn02.place(
    relx=.50,
    rely=.25,
    anchor="center",
)
window.mainloop()
